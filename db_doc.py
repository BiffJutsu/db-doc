'''Script to build a best approximation data dictionary from the INFORMATION_SCHEMA of a given database.
'''
import pyodbc as db
from collections import namedtuple
import openpyxl as ox

Constraint = namedtuple('Constraint',
	['CONSTRAINT_NAME',
	'CONSTRAINT_TYPE',
	'CONSTRAINED_TABLE',
	'CONSTRAINED_COLUMN',
	'SOURCE_TABLE',
	'SOURCE_COLUMN'])

Column = namedtuple('Column',
	['TABLE_NAME',
	'COLUMN_NAME',
	'IS_NULLABLE',
	'DATA_TYPE',
	'MAX_LENGTH',
	'COLUMN_DEFAULT',
	'IS_IDENTITY'])

Config = namedtuple('Config', ['dbstr'])

class Field:
	def __init__(self,
		key=None,
		name=None,
		cap=None,
		typ=None,
		size=None,
		default=None,
		null=None,
		uniq=None,
		ref=None,
		notes=None):
		self.key = key
		self.name = name
		self.cap = cap
		self.typ = typ
		self.size = str(size) if size else size
		self.default = default
		self.null = null
		self.uniq = uniq
		self.ref = ref
		self.notes = notes

	def to_row(self):
		return [
			self.key,
			self.name,
			self.cap,
			self.typ,
			self.size,
			self.default,
			self.null,
			self.uniq,
			self.ref,
			self.notes
		]

class TableDescription:
	def __init__(self, table_name):
		self.headers = ['Primary/Foreign Key','Field Name',
		'Caption','Data Type','Field Size','Default',
		'Nullable', 'Unique', 'Reference','Notes']
		self.table_name = table_name
		self.fields = []

class SchemaCache:
	def __init__(self):
		self.tables = []
		self.columns = []
		self.constraints = []

	@staticmethod
	def bootstrap(dbservice):
		cache = SchemaCache()
		cache.tables = dbservice.get_tables()
		cache.columns = dbservice.get_columns()
		cache.constraints = dbservice.get_constraints()
		dbservice.close()
		return cache

	def get_column_reference(self, table, column):
		refs = []
		for fk in (c for c in self.constraints if c.CONSTRAINT_TYPE == "FOREIGN KEY"):
			if fk.CONSTRAINED_TABLE == table and fk.CONSTRAINED_COLUMN == column:
				refs.append("{0}.{1}".format(fk.SOURCE_TABLE, fk.SOURCE_COLUMN))
		if len(refs) > 0:
			return "; ".join(refs)
		return None

	def is_column_a_key(self, table, column):
		vals = []
		column_constraints = [c
			for c in self.constraints
			if c.CONSTRAINED_TABLE == table and
				c.CONSTRAINED_COLUMN == column]
		if any(c for c in column_constraints if c.CONSTRAINT_TYPE == "PRIMARY KEY"):
			vals.append("P")
		if any(c for c in column_constraints if c.CONSTRAINT_TYPE == "FOREIGN KEY"):
			vals.append("F")
		if len(vals) > 0:
			return "/".join(vals)
		return None

	def is_column_unique(self, table, column):
		return any(c
			for c in self.constraints
			if c.CONSTRAINT_TYPE == "UNIQUE" and
				c.CONSTRAINED_TABLE == table and
				c.CONSTRAINED_COLUMN == column)

	def get_columns_for_table(self, table):
		return (c for c in self.columns if c.TABLE_NAME == table)

class DBService:
	def __init__(self, connstr):
		self.conn = db.connect(connstr)
		self.dbname = self.conn.getinfo(db.SQL_DATABASE_NAME)

	def close(self):
		self.conn.close()

	def get_tables(self):
		sql = """SELECT TABLE_CATALOG, TABLE_NAME
		FROM INFORMATION_SCHEMA.TABLES
		WHERE TABLE_NAME != 'sysdiagrams' AND
		TABLE_TYPE = 'BASE TABLE';"""
		rows = self.conn.cursor().execute(sql).fetchall()
		return [r.TABLE_NAME for r in rows]

	def get_columns(self):
		sql = """SELECT c.TABLE_NAME, c.COLUMN_NAME, c.IS_NULLABLE,
		c.DATA_TYPE, c.CHARACTER_MAXIMUM_LENGTH, c.COLUMN_DEFAULT,
		CAST(COLUMNPROPERTY(object_id(c.TABLE_NAME), c.COLUMN_NAME, 'IsIdentity') as bit) IS_IDENTITY
		FROM INFORMATION_SCHEMA.COLUMNS c
		INNER JOIN INFORMATION_SCHEMA.TABLES t on c.TABLE_NAME = t.TABLE_NAME
		WHERE t.TABLE_NAME != 'sysdiagrams' AND TABLE_TYPE = 'BASE TABLE'
		ORDER BY c.TABLE_NAME, c.ORDINAL_POSITION;"""
		rows = self.conn.cursor().execute(sql).fetchall()
		return [
			Column(r.TABLE_NAME,
				r.COLUMN_NAME,
				r.IS_NULLABLE,
				r.DATA_TYPE,
				r.CHARACTER_MAXIMUM_LENGTH,
				r.COLUMN_DEFAULT,
				r.IS_IDENTITY)
			for r in rows]

	def get_constraints(self):
		sql = """SELECT t.CONSTRAINT_NAME, t.CONSTRAINT_TYPE, t.TABLE_NAME CONSTRAINED_TABLE,
		k.COLUMN_NAME CONSTRAINED_COLUMN, f.TABLE_NAME SOURCE_TABLE, f.COLUMN_NAME SOURCE_COLUMN
		FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS t
		INNER JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE k on t.CONSTRAINT_NAME = k.CONSTRAINT_NAME
		LEFT JOIN INFORMATION_SCHEMA.REFERENTIAL_CONSTRAINTS r on t.CONSTRAINT_NAME = r.CONSTRAINT_NAME
		LEFT JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE f on r.UNIQUE_CONSTRAINT_NAME = f.CONSTRAINT_NAME
		WHERE t.CONSTRAINT_TYPE IN ('PRIMARY KEY', 'FOREIGN KEY', 'UNIQUE') AND
		t.TABLE_NAME != 'sysdiagrams';"""
		rows = self.conn.cursor().execute(sql).fetchall()
		return [
			Constraint(r.CONSTRAINT_NAME,
				r.CONSTRAINT_TYPE,
				r.CONSTRAINED_TABLE,
				r.CONSTRAINED_COLUMN,
				r.SOURCE_TABLE,
				r.SOURCE_COLUMN)
			for r in rows]

class XLService:
	def __init__(self, directory, dbname):
		self.fullpath = "{0}\\{1}_DataDictionary.xlsx".format(directory, dbname)
		self.wkbk = ox.Workbook()
		self.wkbk.remove_sheet(self.wkbk.get_sheet_by_name("Sheet"))

	def _add_sheet(self, table_name):
		self.wkbk.create_sheet(title=table_name)
		return self.wkbk.get_sheet_by_name(table_name)

	def _write_headers(self, ws, headers):
		width = len(headers)
		for j in range(1, width + 1):
			ws.cell(row=1, column=j).value = headers[j-1]

	def _write_fields(self, ws, table_desc):
		row_iterable = ws.iter_rows(min_row=2,
			max_col=len(table_desc.headers),
			max_row=len(table_desc.fields)+1)
		fields = (f.to_row() for f in table_desc.fields)
		for row, field in zip(row_iterable, fields):
			for cell, val in zip(row, field):
				cell.value = val

	def describe_table(self, table_desc):
		wksht = self._add_sheet(table_desc.table_name)
		self._write_headers(wksht, table_desc.headers)
		self._write_fields(wksht, table_desc)

	def save(self):
		self.wkbk.save(self.fullpath)

class DDSpider:
	def __init__(self, dbsvc, cache, xlsvc):
		self.dbservice = dbsvc
		self.cache = cache
		self.xlservice = xlsvc

	def run(self):
		for table in self.cache.tables:
			desc = TableDescription(table)
			for col in self.cache.get_columns_for_table(table):
				field = self.make_field(table, col)
				desc.fields.append(field)
			self.xlservice.describe_table(desc)
		self.xlservice.save()

	def make_field(self, table, col):
		field = Field(name=col.COLUMN_NAME,
			typ=col.DATA_TYPE,
			null=col.IS_NULLABLE,
			size=col.MAX_LENGTH,
			default=col.COLUMN_DEFAULT)
		field.uniq = self.cache.is_column_unique(table, col.COLUMN_NAME)
		field.key = self.cache.is_column_a_key(table, col.COLUMN_NAME)
		field.ref = self.cache.get_column_reference(table, col.COLUMN_NAME)
		if col.IS_IDENTITY:
			field.cap = "Identity"
		return field

def connection_string(server, catalog):
	return 'Driver={SQL Server Native Client 11.0};' + \
	'Server={0};Database={1};'.format(server, catalog) + \
	'Trusted_Connection=yes;APP=DDSpider;'

def discoverdb(server, catalog, directory):
	cstr = connection_string(server, catalog)
	config = Config(dbstr=cstr)
	dbsvc = DBService(config.dbstr)
	cache = SchemaCache.bootstrap(dbsvc)
	xlservice = XLService(directory, dbsvc.dbname)
	spider = DDSpider(dbsvc, cache, xlservice)
	spider.run()
	print("Data Dictionary Created: {0}".format(xlservice.fullpath))

def main():
	server = 'INSERT SERVER'
	catalog = 'INSERT DATABASE'
	directory = r'\\your\filesystem\location\here'
	discoverdb(server, catalog, directory)

if __name__ == '__main__':
	main()
